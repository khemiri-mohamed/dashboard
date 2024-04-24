import csv
import io
import os
import platform
import subprocess
import warnings
from datetime import datetime

from django.db.models import ProtectedError
from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import permissions
from rest_framework import status, generics
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from assets.fct import fctcore

warnings.filterwarnings('ignore', message='Unverified HTTPS request')

class InstanceDetail(generics.GenericAPIView):
    serializer_class = None
    model_class = None
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    parser_classes = (MultiPartParser, FormParser)

    def get_obj(self, pk):
        try:
            return self.model_class.objects.get(pk=pk)
        except:
            return None

    def get(self, request, pk):
        obj = self.get_obj(pk=pk)
        if obj is None:
            return Response({"status": "fail", "message": f"{self.model_class._meta.verbose_name} with Id: {pk} not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.serializer_class(obj)
        return Response({"status": "success", "data": serializer.data})


class InstanceDelete(generics.GenericAPIView):
    serializer_class = None
    model_class = None
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    parser_classes = (MultiPartParser, FormParser)

    def get_obj(self, pk):
        try:
            return self.model_class.objects.get(pk=pk)
        except:
            return None

    def post(self, request):
        pk = request.POST['id']
        obj = self.get_obj(pk)
        if obj is None:
            return Response({"status": "fail", "message": f"{self.model_class._meta.verbose_name} with Id: {pk} not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            obj.delete()
        except ProtectedError as e:
            related_model_name = list(e.protected_objects)[0].__class__._meta.verbose_name
            return Response({"status": "some_relation_exists", "message": 'Cannot delete has'+' '+related_model_name+' '+'exists'}, status=status.HTTP_200_OK)
        except Exception:
            return Response({"status": "fail"}, status=status.HTTP_200_OK)

        return Response({"status": "success"}, status=status.HTTP_200_OK)


class InstanceDeleteSelected(generics.GenericAPIView):
    serializer_class = None
    model_class = None
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    def get_obj(self, pk):
        try:
            return self.model_class.objects.get(pk=pk)
        except:
            return None

    def post(self, request):
        ids = request.data.get('ids', [])
        for pk in ids:
            obj = self.get_obj(pk)
            if obj is None:
                return Response({"status": "fail", "message": f"{self.model_class._meta.verbose_name} with Id: {pk} not found"}, status=status.HTTP_404_NOT_FOUND)

            try:
                obj.delete()
            except ProtectedError as e:
                related_model_name = list(e.protected_objects)[0].__class__._meta.verbose_name
                return Response({"status": "some_relation_exists", "message": 'Cannot delete has'+' '+related_model_name+' '+'exists'}, status=status.HTTP_200_OK)
            except Exception:
                return Response({"status": "fail"}, status=status.HTTP_200_OK)

        return Response({"status": "success"}, status=status.HTTP_200_OK)


class export_csv(generics.GenericAPIView):
    model_class = None
    exclude = None

    def get(self, request):
        columns = [f.name for f in self.model_class._meta.get_fields() if f.name not in self.exclude]

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response['Content-Disposition'] = 'attachment; filename='+self.model_class._meta.verbose_name+'.csv'

        # Create a csv writer object and write the column names to the first row
        writer = csv.writer(response)
        writer.writerow([fctcore.preg_repace(patt='_+', repl=' ', subj=col).strip().title() for col in columns])

        objs = self.model_class.objects.all().order_by('id').values_list(*columns)
        for obj in objs:
            writer.writerow(obj)

        # Return the response object containing the csv file
        return response


class export_xlsx(generics.GenericAPIView):
    model_class = None
    exclude = None

    def get(self, request):
        columns = [f.name for f in self.model_class._meta.get_fields() if f.name not in self.exclude]

        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Add the headers to the worksheet
        row_num = 1
        for col_num, column_title in enumerate([fctcore.preg_repace(patt='_+', repl=' ', subj=col).strip().title() for col in columns], 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title

        objs = self.model_class.objects.all().order_by('id')
        for obj in objs:
            row_num += 1
            row = [getattr(obj, field) for field in columns]
            for col_num, cell_value in enumerate(row, 1):
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.replace(tzinfo=None)
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # create the file response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.dashboardml.sheet')
        response['Content-Disposition'] = 'attachment; filename="'+self.model_class._meta.verbose_name+'.xlsx"'

        # save the workbook to the response
        xlsx_data = io.BytesIO()
        wb.save(xlsx_data)
        xlsx_data.seek(0)
        response.write(xlsx_data.read())
        return response


class process_script:
    def __init__(self, script_name):
        self.script_name = script_name
        self.project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.manage_py_path = os.path.join(self.project_path, 'manage.py')
        self.platform_sys = platform.system()
        self.pid_file = 'pid.txt'

        if self.platform_sys == 'Windows':
            self.python = 'python'
        else:
            self.python = 'python3'

    def run(self):
        process = subprocess.Popen([self.python, self.manage_py_path, 'runscript', self.script_name])
        with open(self.pid_file, 'w') as pid_file:
            pid_file.write(str(process.pid))

    def stop(self):
        with open(self.pid_file, 'r') as f:
            pid = int(f.read().strip())
            if self.platform_sys == 'Windows':
                os.system(f'taskkill /F /PID {pid}')
            else:
                os.system(f'kill -9 {pid}')
